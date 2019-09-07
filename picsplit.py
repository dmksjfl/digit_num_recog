# -*- encoding:utf-8 -*-

import cv2
import numpy as np
from PIL import Image
from matplotlib import pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Fill

def split():
  rawim = cv2.imread('d:/test/pic/player_num3.jpg')
  #denoising
  rawimg = cv2.fastNlMeansDenoising(rawim, None, 3, 21, 7)
  fig = plt.figure()
  #use plt to show pictures
  fig.add_subplot(2, 3, 1)
  plt.title("raw image")
  plt.imshow(rawimg)

  fig.add_subplot(2, 3, 2)
  plt.title("grey scale image")
  grayscaleimg = cv2.cvtColor(rawimg, cv2.COLOR_BGR2GRAY)
  grayscaleimg = grayscaleimg - int(np.mean(grayscaleimg))
  grayscaleimg[grayscaleimg < 110] = 0
  ret,grayscaleimg=cv2.threshold(grayscaleimg,0,255,cv2.THRESH_BINARY)
  grayscaleimg = 255- grayscaleimg
  #start erosion and repeat 5 times
  kernel = np.ones((6,6),np.uint8)
  erosion = cv2.erode(grayscaleimg,kernel,iterations = 3)
  #dilation 3 times to ensure good split
  dilation = cv2.dilate(erosion,kernel,iterations = 3)
  grayscaleimg = dilation
  #print numpy array of grayscaleimg
  #print(grayscaleimg)
  cv2.imwrite('d:/test/pic/gray.jpg', grayscaleimg)

  plt.imshow(grayscaleimg, cmap='gray')
  #show erosion result
  fig.add_subplot(2,3,3)
  plt.imshow(erosion)
  #show dilation result
  fig.add_subplot(2,3,6)
  plt.imshow(dilation)
  
  # counting non-zero value by row , axis y
  row_nz = []
  for row in grayscaleimg.tolist():
      row_nz.append(len(row) - row.count(0))
  #show row_nz
  #print(row_nz)
  
  # counting non-zero value by column, x axis
  col_nz = []
  for col in grayscaleimg.T.tolist():
      col_nz.append(len(col) - col.count(0))
  #show col_nz
  #print(col_nz)

  # start split
  # first find upper and lower boundary of y (row)
  fig.add_subplot(2, 3, 5)
  plt.title("sliced image")
  upper_y = 0
  for i, x in enumerate(row_nz):
      if x != 0:
          upper_y = i
          break
  lower_y = 0
  for i, x in enumerate(row_nz[::-1]):
      if x != 0:
          lower_y = len(row_nz) - i
          break
  sliced_y_img = rawim[upper_y:lower_y, :]
  #there is actually no need to save this image, you can disable it
  cv2.imwrite('d:/test/pic/slicedimg.jpg', sliced_y_img)
  plt.imshow(sliced_y_img)

  # then we find left and right boundary of every digital (x, on column)
  column_boundary_list = []
  for i, x in enumerate(col_nz[:-1]):
      if (col_nz[i] == 0 and col_nz[i + 1] != 0) or col_nz[i] != 0 and col_nz[i + 1] == 0:
          column_boundary_list.append(i + 1)
  img_list = []
  xl = [column_boundary_list[i:i + 2] for i in range(0, len(column_boundary_list), 2)]
  for x in xl:
      img_list.append(sliced_y_img[:, x[0]:x[1]])
  # del invalid image
  img_list = [x for x in img_list if x.shape[1] < 100]
  # show image
  fig = plt.figure()
  for i, imga in enumerate(img_list[:len(img_list)]):
      fig.add_subplot(3, 4, i + 1)
      plt.imshow(imga)
      plt.imsave('d:/test/pic/%s.jpg' % i, imga)
  plt.show()

  # write the sliced number picture into excel
  outDir = 'd:/test/pic/'
  for i in range(0, len(img_list)-1):
      image = Image.open('d:/test/pic/' + str(i) + '.jpg')
      wb = openpyxl.Workbook()  # creat Excel file
      sheet = wb.create_sheet("cutimage")  # creat a sheet
      imgW, imgH = image.size  # acquire the size of the picture
      for w in range(imgW):
          for h in range(imgH):
              # fill each pixel with the color of the corresponding cell background color
              rgba = image.getpixel((w, h))
              colorHex = hex(rgba[0])[2:].zfill(2) + hex(rgba[1])[2:].zfill(2) + hex(rgba[2])[2:].zfill(2)
              fill = PatternFill(fill_type='solid', start_color=colorHex, end_color=colorHex)
              sheet.cell(row=h + 1, column=w + 1).fill = fill
      wb.save(outDir + str(i) + '.xlsx')  # save xlsx file

split()